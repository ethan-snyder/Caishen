"""
bonds.py — bonds_info()

Fixed-income data across three layers:

  1. U.S. Treasury curve -- the Treasury's own daily par yield curve CSV
     (`_us_treasury_curve`), which is the authoritative source and covers
     every requested tenor (1-MO through 30-YR) in one fetch. Falls back
     to yfinance's ^IRX/^FVX/^TNX/^TYX (3-MO/5-YR/10-YR/30-YR only) if
     Treasury.gov is unreachable. Daily, live.
  2. Foreign sovereign yields -- as rich a curve as a genuinely free, live
     source publishes per country (see FOREIGN_SOVEREIGNS below); the
     10-year baseline is OECD's "Long-Term Government Bond Yields:
     10-Year" series via FRED, which is MONTHLY and published with a lag
     (typically ~1-2 months) -- every sovereign row carries its own
     observation date so a lagged figure is never presented as today's
     quote.
  3. Corporate bonds -- ICE BofA index effective yields by credit rating
     via FRED. Daily, close.

On corporate "issuers" specifically: individual corporate bond quotes
(Apple's 2032s at X%) have no free, reliable data source -- that's a
subscription-data product. What IS free, and is what those issuers
actually price against, is the ICE BofA rating-tier benchmark curve. So
the two corporate sections are organized by rating tier (AAA/AA/A/BBB and
BB/B/CCC) rather than by named issuer, which is real data rather than
fabricated per-company yields.
"""

from concurrent.futures import ThreadPoolExecutor

import yfinance as yf

from utils import fmt_pct_raw
from logger import log_event, warn

# yfinance's ^IRX/^FVX/^TNX/^TYX already report the yield directly as a
# percent (e.g. a close of 4.35 means 4.35%), same convention as
# dividendYield -- so these use fmt_pct_raw(), not fmt_pct(). Used as a
# fallback if Treasury.gov's own curve (see below) is unreachable.
TREASURIES = {
    "13-Week T-Bill": "^IRX",
    "5-Year Treasury": "^FVX",
    "10-Year Treasury": "^TNX",
    "30-Year Treasury": "^TYX",
}

TREASURY_TERMS = {
    "13-Week T-Bill": "3-MO",
    "5-Year Treasury": "5-YR",
    "10-Year Treasury": "10-YR",
    "30-Year Treasury": "30-YR",
}

# U.S. Treasury's own daily par yield curve -- the primary/authoritative
# source, since it's the one place all seven requested tenors (1-MO
# through 30-YR) are published together, daily, for free, no key needed.
# https://home.treasury.gov/resource-center/data-chart-center/interest-rates
US_TREASURY_CSV_URL = (
    "https://home.treasury.gov/resource-center/data-chart-center/"
    "interest-rates/daily-treasury-rates.csv/{year}/all"
    "?type=daily_treasury_yield_curve&field_tdr_date_value={year}"
)
# (our term code, display name, Treasury.gov's CSV column header)
US_TREASURY_CURVE_TENORS = [
    ("1-MO", "1-Month Treasury", "1 Mo"),
    ("3-MO", "13-Week T-Bill", "3 Mo"),
    ("6-MO", "6-Month Treasury", "6 Mo"),
    ("1-YR", "1-Year Treasury", "1 Yr"),
    ("5-YR", "5-Year Treasury", "5 Yr"),
    ("10-YR", "10-Year Treasury", "10 Yr"),
    ("30-YR", "30-Year Treasury", "30 Yr"),
]


def _us_treasury_curve():
    """
    Parses Treasury.gov's daily par yield curve CSV (newest date first).
    Returns a list of {term, name, yield, change, prev} covering every
    tenor in US_TREASURY_CURVE_TENORS, using the latest row and, for
    change, the row before it. Raises on any structural surprise so the
    caller can fall back to the yfinance-based curve instead of shipping
    a partially-wrong row.
    """
    import csv
    import datetime
    import io

    import requests

    year = datetime.date.today().year
    resp = requests.get(US_TREASURY_CSV_URL.format(year=year), timeout=10)
    resp.raise_for_status()
    rows = list(csv.DictReader(io.StringIO(resp.text)))
    if not rows:
        raise ValueError("no rows returned from Treasury.gov")

    def parse(row):
        out = {}
        for term, _name, col in US_TREASURY_CURVE_TENORS:
            raw = (row.get(col) or "").strip()
            out[term] = float(raw) if raw not in ("", "N/A") else None
        return out

    latest, prior = parse(rows[0]), (parse(rows[1]) if len(rows) > 1 else None)

    treasuries = []
    for term, name, _col in US_TREASURY_CURVE_TENORS:
        last = latest.get(term)
        prev = prior.get(term) if prior else None
        treasuries.append({
            "term": term, "name": name, "yield": last,
            "change": (last - prev) if last is not None and prev is not None else None,
            "prev": prev,
        })
    return treasuries

# OECD 10-year government bond yields via FRED. Series ID pattern is
# IRLTLT01<ISO2>M156N (M = monthly, 156N = percent per annum).
#
# JGB and GILT ship visible (the two get an enlarged "hero" tile with a
# mini yield curve on the Bonds page); the rest are available in the
# page's edit-layout tray. Brazil isn't an OECD member (it's a key
# partner), so its series coverage is less certain than the others -- it
# degrades to N/A rather than breaking the section if FRED doesn't have it.
#
# The goal on every row is the fullest curve (1-MO/3-MO/6-MO/1-YR/5-YR/
# 10-YR/30-YR) a genuinely free, live source supports for that country --
# no tenor is ever interpolated or invented to fill a gap. In practice:
#
#   - JGB: Japan's Ministry of Finance -- the issuer itself -- publishes a
#     real daily par-yield curve across every tenor from 1Y to 40Y as a
#     plain CSV (`_jgb_curve_from_mof`). Real 1Y/5Y/10Y/30Y.
#   - CANGB: the Bank of Canada's Valet API publishes real daily benchmark
#     bond yields (`_boc_curve`). Real 5Y/10Y/30Y.
#   - GILT: the UK's DMO stopped publishing daily gilt reference prices in
#     2017 (that moved to FTSE-Tradeweb, which isn't free). The Bank of
#     England does independently publish its own daily fitted gilt curve
#     with a real 1-MO/6-MO/1Y/5Y/30Y, but only as a ZIP of Excel
#     workbooks with an undocumented layout, and the workbook bundles
#     *multiple* curve types (nominal/real/inflation/OIS) across several
#     sheets. `_gilt_curve_from_boe`/`_scan_boe_sheet` below implement a
#     generic column scanner for it, but it is NOT currently wired into
#     GILT's spec (see `boe_curve` note) -- a real run against the actual
#     file surfaced it matching the wrong sheet/curve type for some
#     tenors, which is worse than missing data (a plausible-looking but
#     wrong basis-point figure), and this sandbox has no outbound route
#     to the file to fix the scan against ground truth. GILT stays on the
#     verified-safe 3-MO/10-YR pair below until that can be confirmed
#     against the real workbook.
#   - All six also get a `short_series` 3-month point layered on: GILT's
#     is OECD's actual UK Treasury-bill series; the rest use the OECD
#     3-month interbank rate, the closest published short rate for that
#     country (a market-rate proxy, not a government-bill yield).
#   - No free, live, unauthenticated source was found for 1-MO/6-MO/1-YR
#     for BUND/ACGB, or for 1-MO/6-MO for CANGB/JGB -- Germany's
#     Bundesbank and Australia's RBA publish curves but not through a
#     simple stable free endpoint verified here. Those tenors are left off
#     rather than invented -- each row's `curve` only ever contains real
#     points.
FOREIGN_SOVEREIGNS = [
    {"key": "jgb",   "name": "Japan 10Y",     "label": "JGB",   "series": "IRLTLT01JPM156N", "default_hidden": False,
     "mof_curve": True, "short_series": "IR3TIB01JPM156N", "short_label": "3-MO"},
    {"key": "gilt",  "name": "U.K. 10Y",      "label": "GILT",  "series": "IRLTLT01GBM156N", "default_hidden": False,
     "short_series": "IR3TTS01GBM156N", "short_label": "3-MO"},
    {"key": "bund",  "name": "Germany 10Y",   "label": "BUND",  "series": "IRLTLT01DEM156N", "default_hidden": True,
     "short_series": "IR3TIB01DEM156N", "short_label": "3-MO"},
    {"key": "cangb", "name": "Canada 10Y",    "label": "CAN",   "series": "IRLTLT01CAM156N", "default_hidden": True,
     "boc_curve": True, "short_series": "IR3TIB01CAM156N", "short_label": "3-MO"},
    {"key": "acgb",  "name": "Australia 10Y", "label": "ACGB",  "series": "IRLTLT01AUM156N", "default_hidden": True,
     "short_series": "IR3TIB01AUM156N", "short_label": "3-MO"},
    {"key": "ntnb",  "name": "Brazil 10Y",    "label": "BRA",   "series": "IRLTLT01BRM156N", "default_hidden": True},
]

# Sovereign tiles that get the enlarged "hero" treatment on the Bonds page.
HERO_SOVEREIGNS = {"jgb", "gilt"}

# Japan MOF's real daily par-yield curve (1Y-40Y). Picks the tenors the
# hero tile displays -- the raw file has every tenor if more are wanted.
JGB_CSV_URL = "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/jgbcme.csv"
JGB_CURVE_TENORS = ["1Y", "5Y", "10Y", "30Y"]


def _jgb_curve_from_mof():
    """
    Parses MOF's "current data" CSV: a title line, a header row
    (`Date,1Y,2Y,...,40Y`), daily rows, then a blank line and a footer
    note -- both of which get skipped since they don't parse as a date
    row. Returns (curve_points, ten_year_yield, ten_year_change) using
    the latest row and, for change, the prior one.
    """
    import csv
    import io
    import requests

    resp = requests.get(JGB_CSV_URL, timeout=10)
    resp.raise_for_status()
    rows = list(csv.reader(io.StringIO(resp.text)))

    header_idx = next(i for i, r in enumerate(rows) if r and r[0].strip() == "Date")
    header = [c.strip() for c in rows[header_idx]]

    data_rows = []
    for r in rows[header_idx + 1:]:
        if not r or not r[0].strip():
            continue
        try:
            values = [float(v) for v in r[1:len(header)]]
        except ValueError:
            continue
        data_rows.append(dict(zip(header[1:], values)))

    if not data_rows:
        raise ValueError("no JGB curve rows parsed from MOF CSV")

    latest = data_rows[-1]
    prior = data_rows[-2] if len(data_rows) >= 2 else None

    curve = [
        {"tenor": t, "yield": latest[t]}
        for t in JGB_CURVE_TENORS if t in latest
    ]
    ten_year = latest.get("10Y")
    ten_year_change = (
        ten_year - prior["10Y"]
        if prior and ten_year is not None and "10Y" in prior else None
    )
    return curve, ten_year, ten_year_change


# Bank of Canada Valet API -- real daily benchmark Government of Canada
# bond yields, no key required. https://www.bankofcanada.ca/valet/docs
BOC_OBSERVATIONS_URL = "https://www.bankofcanada.ca/valet/observations/{series}/json?recent=2"
BOC_CURVE_SERIES = [
    ("5Y", "BD.CDN.5YR.DQ.YLD"),
    ("10Y", "BD.CDN.10YR.DQ.YLD"),
    ("30Y", "BD.CDN.LONG.DQ.YLD"),
]


def _boc_curve():
    """
    Real Government of Canada benchmark bond yields (5Y/10Y/30Y) straight
    from the Bank of Canada's own Valet API. Returns (curve_points,
    ten_year_yield, ten_year_change, date) using the latest observation
    and, for change, the prior one.
    """
    import requests

    series_ids = ",".join(s for _, s in BOC_CURVE_SERIES)
    resp = requests.get(BOC_OBSERVATIONS_URL.format(series=series_ids), timeout=10)
    resp.raise_for_status()
    obs = resp.json().get("observations") or []
    if not obs:
        raise ValueError("no observations returned from Bank of Canada Valet API")

    latest, prior = obs[0], (obs[1] if len(obs) > 1 else None)

    def _val(row, series_id):
        cell = row.get(series_id)
        return float(cell["v"]) if cell and cell.get("v") not in (None, "") else None

    curve = []
    for tenor, series_id in BOC_CURVE_SERIES:
        v = _val(latest, series_id)
        if v is not None:
            curve.append({"tenor": tenor, "yield": v})

    ten_year = _val(latest, "BD.CDN.10YR.DQ.YLD")
    ten_year_change = None
    if prior and ten_year is not None:
        prior_ten_year = _val(prior, "BD.CDN.10YR.DQ.YLD")
        if prior_ten_year is not None:
            ten_year_change = ten_year - prior_ten_year

    return curve, ten_year, ten_year_change, latest.get("d")


# Bank of England's daily gilt nominal yield curve -- the only free,
# official, live source for GILT tenors beyond 3-MO/10-YR (the UK DMO
# stopped publishing daily reference prices in 2017). BoE doesn't expose
# this over an API -- only as a ZIP of Excel workbooks, refreshed daily.
# https://www.bankofengland.co.uk/statistics/yield-curves
#
# NOT CURRENTLY WIRED UP. This sandbox has no outbound route to actually
# download the file, so the scanner below was built and tested only
# against a synthetic workbook constructed from BoE's *written*
# description of the layout -- not the real one. When it was briefly
# enabled against the live file, it returned only a partial/inconsistent
# match (the real workbook bundles multiple curve types -- nominal, real,
# inflation, OIS -- across several sheets, and the generic scan can't yet
# tell them apart, so it risks silently attributing a real/OIS-curve
# point to the nominal gilt tile). That's worse than the gap it was meant
# to fill, so GILT stays on the verified-safe 3-MO/10-YR pair
# (`short_series` + the FRED 10Y headline) until someone with real
# network access to bankofengland.co.uk can confirm the actual sheet
# names/columns and this gets fixed against ground truth. Left in place,
# unused, as a starting point for that fix rather than deleted.
BOE_YIELD_CURVE_ZIP_URL = "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip"
# (tenor label, maturity in years) -- BoE's published header row uses
# years-to-maturity as numeric column labels (per their own FAQ: "the
# short end ... in granular monthly intervals, [and] the whole curve ...
# in semi-annual intervals"), so each wanted tenor is matched to the
# closest header value within BOE_TENOR_TOLERANCE years.
BOE_CURVE_TENORS = [("1-MO", 1 / 12), ("6-MO", 0.5), ("1Y", 1.0), ("5Y", 5.0), ("30Y", 30.0)]
BOE_TENOR_TOLERANCE = 0.02


def _gilt_curve_from_boe():
    """
    UNUSED -- see the NOT CURRENTLY WIRED UP note above BOE_YIELD_CURVE_ZIP_URL.

    Downloads BoE's published yield-curve ZIP and scans every worksheet
    in every workbook inside it for a header row of numeric years-to-
    maturity, matching the closest column to each tenor in
    BOE_CURVE_TENORS and reading the last populated data row below it
    (the most recent date). This is a generic scan rather than hardcoded
    cell references, since BoE doesn't document the exact layout -- but
    it's still a best-effort match against an unverified file structure
    (this sandbox has no route to actually download and inspect the file
    ahead of time). Any point it can't confidently match is just left
    out of the returned curve rather than guessed; a total failure here
    is caught by the caller and GILT simply keeps its existing 3-MO/10-YR
    curve, so a layout change on BoE's end degrades gracefully rather
    than shipping a silently-misread number. In practice this alone
    wasn't enough -- see the note above for why it's disabled.
    """
    import io
    import zipfile

    import openpyxl
    import requests

    resp = requests.get(BOE_YIELD_CURVE_ZIP_URL, timeout=20)
    resp.raise_for_status()

    found = {}
    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        for name in zf.namelist():
            if not name.lower().endswith((".xlsx", ".xls")):
                continue
            with zf.open(name) as fh:
                wb = openpyxl.load_workbook(io.BytesIO(fh.read()), read_only=True, data_only=True)
                for ws in wb.worksheets:
                    _scan_boe_sheet(ws, found)
                wb.close()

    if not found:
        raise ValueError("couldn't locate a recognizable maturity column in the BoE workbook")

    order = dict(BOE_CURVE_TENORS)
    curve = sorted(
        ({"tenor": label, "yield": v} for label, v in found.items()),
        key=lambda p: order.get(p["tenor"], 0),
    )
    return curve


def _scan_boe_sheet(ws, found):
    """Looks at each sheet's first 15 rows for one that's mostly numeric
    (a maturity-in-years header), matches unfound tenors to the closest
    column within tolerance, then reads the latest data row under it."""
    max_col = min(ws.max_column or 0, 400)
    for row_idx in range(1, 16):
        header = []
        for col in range(1, max_col + 1):
            v = ws.cell(row=row_idx, column=col).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                header.append((col, float(v)))
        if len(header) < 5:
            continue

        col_for_tenor = {}
        for label, years in BOE_CURVE_TENORS:
            if label in found:
                continue
            closest_col, closest_years = min(header, key=lambda ch: abs(ch[1] - years))
            if abs(closest_years - years) <= BOE_TENOR_TOLERANCE:
                col_for_tenor[label] = closest_col
        if not col_for_tenor:
            continue

        any_col = next(iter(col_for_tenor.values()))
        last_row = None
        for r in range(row_idx + 1, (ws.max_row or row_idx) + 1):
            v = ws.cell(row=r, column=any_col).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                last_row = r
        if last_row is None:
            continue

        for label, col in col_for_tenor.items():
            v = ws.cell(row=last_row, column=col).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                found[label] = float(v)


# ICE BofA effective yields by rating tier (daily, close).
CORPORATE_IG = [
    {"key": "aaa", "label": "AAA", "name": "Aaa/AAA", "series": "BAMLC0A1CAAAEY"},
    {"key": "aa",  "label": "AA",  "name": "Aa/AA",   "series": "BAMLC0A2CAAEY"},
    {"key": "a",   "label": "A",   "name": "A",       "series": "BAMLC0A3CAEY"},
    {"key": "bbb", "label": "BBB", "name": "Baa/BBB", "series": "BAMLC0A4CBBBEY"},
]

CORPORATE_HY = [
    {"key": "bb",  "label": "BB",  "name": "Ba/BB",       "series": "BAMLH0A1HYBBEY"},
    {"key": "b",   "label": "B",   "name": "B",           "series": "BAMLH0A2HYBEY"},
    {"key": "ccc", "label": "CCC", "name": "Caa/CCC & <", "series": "BAMLH0A3HYCEY"},
]


def _get_bond_data():
    results = {}
    for name, symbol in TREASURIES.items():
        try:
            hist = yf.Ticker(symbol).history(period="5d")
            if hist.empty:
                raise ValueError("no data returned")
            yield_pct = float(hist["Close"].iloc[-1])
            results[name] = yield_pct
        except Exception as e:
            warn(f"Couldn't fetch {name} ({e})")
            results[name] = None
    return results


def bonds_info():
    log_event("Fetching Treasury yield data")
    print("\n===== U.S. Treasury Yields =====")
    data = _get_bond_data()
    for name, yield_pct in data.items():
        print(f"  {name:<20} {fmt_pct_raw(yield_pct)}")
    print("=" * 40)
    print()
    log_event("Treasury yield data fetched")
    return data


# ----------------------------------------------------------------------
# FRED-backed rows (foreign sovereigns + corporate rating tiers)
# ----------------------------------------------------------------------

def _fred_row(spec):
    """
    {latest yield, change vs. prior observation, observation date, curve}
    for one sovereign or corporate row.

    The headline 10Y figure and the `curve` list come from whichever
    primary source is real and richest for that country: Japan's MOF for
    JGB (`mof_curve`), the Bank of Canada's Valet API for CANGB
    (`boc_curve`), and plain FRED/OECD for everything else (corporate
    rows included -- they never set either flag, so they just take the
    FRED branch with no `short_series` and end up with `curve: None`).

    A `short_series` (3-month OECD rate) is layered on top of whichever
    primary curve was built, so most sovereigns end up with a genuine
    two-plus-point curve rather than a single 10Y level.

    Every failure mode (no API key, unknown series, empty series, source
    unreachable) degrades gracefully -- a failed primary source just
    leaves `yield`/`curve` as None/empty rather than raising, and a
    failed short-end fetch just skips that one point.
    """
    row = {
        "key": spec["key"],
        "label": spec.get("label", spec["key"].upper()),
        "name": spec.get("name", spec["key"]),
        "yield": None, "change": None, "date": None,
        "default_hidden": bool(spec.get("default_hidden", False)),
        "curve": None,
    }

    curve_points = []

    if spec.get("mof_curve"):
        try:
            curve_points, row["yield"], row["change"] = _jgb_curve_from_mof()
            # MOF's file doesn't carry a per-row date we parse out --
            # leave `date` unset rather than guess; the tile only shows
            # `date` for FRED-sourced (monthly, lagged) rows anyway.
        except Exception as e:
            warn(f"Couldn't fetch JGB curve from MOF ({e})")

    elif spec.get("boc_curve"):
        try:
            curve_points, row["yield"], row["change"], row["date"] = _boc_curve()
        except Exception as e:
            warn(f"Couldn't fetch CANGB curve from Bank of Canada ({e})")

    else:
        try:
            # Imported lazily and by name: bonds.py doesn't otherwise
            # depend on market_info, and importing at module scope would
            # make a FRED problem in one module surface as an import
            # error in the other.
            import market_info
            series = market_info._fred_series_latest_and_history(spec["series"], points=2)
            if not series:
                raise ValueError("no observations returned")
            row["yield"] = series["value"]
            row["date"] = series["date"]
            history = series.get("history") or []
            if len(history) >= 2:
                row["change"] = series["value"] - history[-2]["value"]
        except Exception as e:
            warn(f"Couldn't fetch {spec.get('name', spec['key'])} [{spec['series']}] ({e})")

    # `boe_curve` is intentionally not read here -- see the note above
    # BOE_YIELD_CURVE_ZIP_URL for why that source isn't wired up yet.

    short_series = spec.get("short_series")
    if short_series:
        try:
            import market_info
            short = market_info._fred_series_latest_and_history(short_series, points=1)
            if short:
                curve_points = [{"tenor": spec.get("short_label", "3-MO"), "yield": short["value"]}] + curve_points
        except Exception as e:
            warn(f"Couldn't fetch short-end {spec.get('name', spec['key'])} [{short_series}] ({e})")

    # Plain-FRED sovereigns (no mof_curve/boc_curve) only have a headline
    # 10Y so far -- fold it into the curve too, unless the short-end fetch
    # above also failed and left nothing to chart.
    if not any(p["tenor"] in ("10Y", "10-YR") for p in curve_points) and row["yield"] is not None:
        curve_points.append({"tenor": "10-YR", "yield": row["yield"]})

    curve_points.sort(key=lambda p: _tenor_sort_key(p["tenor"]))
    row["curve"] = curve_points or None
    return row


def _tenor_sort_key(label):
    """Maps the various tenor-label formats used across sources ('3-MO',
    '1Y', '10-YR', ...) to years-to-maturity, purely so a tile's curve
    bars render left-to-right in maturity order regardless of which
    source(s) contributed which point."""
    l = label.upper().replace("-", "")
    try:
        if l.endswith("MO"):
            return float(l[:-2]) / 12
        if l.endswith("YR"):
            return float(l[:-2])
        if l.endswith("Y"):
            return float(l[:-1])
    except ValueError:
        pass
    return 999.0


def _fred_rows(specs):
    """Fetches a batch of FRED series in parallel -- 13 sequential round
    trips noticeably delayed the whole /api/bonds response."""
    with ThreadPoolExecutor(max_workers=6) as pool:
        return list(pool.map(_fred_row, specs))


def _us_treasury_curve_yfinance_fallback():
    """The old 4-tenor yfinance-only curve, used only if Treasury.gov
    itself is unreachable -- real data, just narrower coverage."""
    treasuries = []
    for name, symbol in TREASURIES.items():
        try:
            hist = yf.Ticker(symbol).history(period="5d")
            if len(hist) < 2:
                raise ValueError("insufficient history")
            last = float(hist["Close"].iloc[-1])
            prev = float(hist["Close"].iloc[-2])
            treasuries.append({
                "term": TREASURY_TERMS.get(name, name), "name": name,
                "yield": last, "change": last - prev, "prev": prev,
            })
        except Exception as e:
            warn(f"Couldn't fetch {name} ({e})")
            treasuries.append({
                "term": TREASURY_TERMS.get(name, name), "name": name,
                "yield": None, "change": None, "prev": None,
            })
    return treasuries


def get_bond_data_full():
    """
    Data-only variant for API consumers:

      treasuries   -- U.S. curve, 1-MO through 30-YR (Treasury.gov, daily/
                      live; falls back to a narrower yfinance-based curve
                      if Treasury.gov itself is unreachable)
      sovereigns   -- foreign government yields, as many tenors as a real
                      free source publishes per country (see
                      FOREIGN_SOVEREIGNS), 10Y baseline via FRED/OECD
                      (monthly, each row carrying its own observation date)
      corporate_ig -- investment-grade rating tiers (ICE BofA, daily)
      corporate_hy -- high-yield/speculative rating tiers (ICE BofA, daily)
    """
    log_event("Fetching bond data (full)")

    try:
        treasuries = _us_treasury_curve()
    except Exception as e:
        warn(f"Treasury.gov par yield curve unavailable, falling back to yfinance ({e})")
        treasuries = _us_treasury_curve_yfinance_fallback()

    return {
        "treasuries": treasuries,
        "sovereigns": _fred_rows(FOREIGN_SOVEREIGNS),
        "corporate_ig": _fred_rows(CORPORATE_IG),
        "corporate_hy": _fred_rows(CORPORATE_HY),
    }


if __name__ == "__main__":
    bonds_info()
